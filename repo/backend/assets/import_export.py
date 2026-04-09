"""
assets/import_export.py

Core bulk import and export logic, fully decoupled from HTTP concerns.

Import flow (3 phases):
  1. parse_and_classify(file_path, tenant, site)
     → Parses CSV/XLSX, validates headers/rows, returns per-row classification.
  2. apply_corrections(results_json, corrections, tenant, site)
     → Patches field values in REJECTED rows and re-classifies them.
  3. confirm_import(job, decisions, user)
     → Atomically creates/updates assets based on user decisions.

Row statuses:
  NEW              — asset_code not seen in this site; ready to create.
  UPDATE_CANDIDATE — asset_code exists but fingerprint differs; offers update.
  DUPLICATE        — exact fingerprint match; already exists unchanged.
  REJECTED         — format/validation error; must correct before proceeding.
  BATCH_DUPLICATE  — same asset_code appears earlier in the same batch.

Export:
  build_export(user, site_id, format_type) → (bytes, content_type, filename)
"""
import csv
import hashlib
import io
import os
import re
from datetime import timezone as dt_timezone
from pathlib import Path

from django.conf import settings
from django.db import transaction
from django.utils import timezone

from assets.models import Asset, AssetClassification, AssetVersion, BulkImportJob

_ASSET_CODE_RE = re.compile(r"^[A-Z0-9\-]{3,50}$")

# ---------------------------------------------------------------------------
# File parsing
# ---------------------------------------------------------------------------

REQUIRED_HEADERS = {"asset_code", "name", "classification_code"}


def parse_file(file_path: str) -> tuple[list[str], list[dict]]:
    """
    Parse a CSV or XLSX file.
    Returns (normalised_headers, rows) where each row is a plain dict keyed
    by the normalised (lowercased, stripped) header names.
    Row dicts include a synthetic '_row_number' key (1-based, header excluded).
    Raises ValueError on unsupported extension or missing required headers.
    """
    ext = Path(file_path).suffix.lower()
    if ext == ".csv":
        headers, rows = _parse_csv(file_path)
    elif ext in (".xlsx",):
        headers, rows = _parse_xlsx(file_path)
    else:
        raise ValueError(f"Unsupported file format '{ext}'. Only CSV and XLSX are accepted.")

    missing = REQUIRED_HEADERS - set(headers)
    if missing:
        raise ValueError(f"Missing required columns: {', '.join(sorted(missing))}")

    return headers, rows


def _parse_csv(file_path: str) -> tuple[list[str], list[dict]]:
    with open(file_path, newline="", encoding="utf-8-sig") as fh:
        reader = csv.DictReader(fh)
        raw_headers = [h.strip().lower() for h in (reader.fieldnames or [])]
        rows = []
        for i, row in enumerate(reader, start=1):
            normalised = {k.strip().lower(): (v or "").strip() for k, v in row.items()}
            normalised["_row_number"] = i
            rows.append(normalised)
    return raw_headers, rows


def _parse_xlsx(file_path: str) -> tuple[list[str], list[dict]]:
    from openpyxl import load_workbook

    wb = load_workbook(file_path, read_only=True, data_only=True)
    ws = wb.active
    headers: list[str] = []
    rows: list[dict] = []

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i == 0:
            headers = [str(c or "").strip().lower() for c in row]
        else:
            values = [str(c) if c is not None else "" for c in row]
            if not any(values):
                continue
            d = dict(zip(headers, values))
            d["_row_number"] = i  # i is 1-based after header row
            rows.append(d)

    wb.close()
    return headers, rows


# ---------------------------------------------------------------------------
# Row classification
# ---------------------------------------------------------------------------

def parse_and_classify(file_path: str, tenant, site) -> dict:
    """
    Entry point for phase 1. Parses the file and classifies every row.
    Returns the full results_json structure (same shape stored on the job).
    Raises ValueError if headers are invalid.
    """
    headers, rows = parse_file(file_path)

    # Pre-load all classifications for this tenant once (avoids N+1)
    cls_map: dict[str, AssetClassification] = {
        c.code: c
        for c in AssetClassification.objects.filter(tenant=tenant, is_active=True)
    }

    classified = []
    seen_codes: dict[str, int] = {}  # asset_code → first row_number

    for row in rows:
        result = _classify_single_row(row, site, cls_map, seen_codes)
        # Only register non-duplicate, non-rejected codes in seen_codes
        if result["status"] not in ("REJECTED", "BATCH_DUPLICATE") and result["asset_code"]:
            seen_codes.setdefault(result["asset_code"], result["row_number"])
        classified.append(result)

    return {"rows": classified}


def _classify_single_row(
    row: dict,
    site,
    cls_map: dict,
    seen_codes: dict,
) -> dict:
    """Classify one row. Mutates seen_codes as a side-effect for valid rows."""
    row_num = row["_row_number"]
    asset_code     = row.get("asset_code", "").strip().upper()
    name           = row.get("name", "").strip()
    cls_code       = row.get("classification_code", "").strip().upper()
    custom_data    = {
        k: v
        for k, v in row.items()
        if k not in ("asset_code", "name", "classification_code", "_row_number")
        and v != ""
    }

    result: dict = {
        "row_number": row_num,
        "status": "NEW",
        "asset_code": asset_code,
        "name": name,
        "classification_code": cls_code,
        "custom_data": custom_data,
        "errors": [],
        "existing_asset_id": None,
    }

    # ---- Format validation ------------------------------------------------
    errors: list[str] = []
    if not asset_code:
        errors.append("asset_code is required.")
    elif not _ASSET_CODE_RE.match(asset_code):
        errors.append(
            f"asset_code '{asset_code}' must match ^[A-Z0-9\\-]{{3,50}}$."
        )

    if not name:
        errors.append("name is required.")
    elif len(name) > 200:
        errors.append("name must be ≤ 200 characters.")

    if not cls_code:
        errors.append("classification_code is required.")
    elif cls_code not in cls_map:
        errors.append(f"classification_code '{cls_code}' not found in this tenant.")

    if errors:
        result["status"] = "REJECTED"
        result["errors"] = errors
        return result

    # ---- Intra-batch dedup -----------------------------------------------
    if asset_code in seen_codes:
        result["status"] = "BATCH_DUPLICATE"
        result["errors"] = [
            f"Duplicate asset_code in batch "
            f"(first seen on row {seen_codes[asset_code]})."
        ]
        return result

    # ---- Fingerprint & DB matching ----------------------------------------
    cls_obj = cls_map[cls_code]
    raw = "|".join([str(site.pk), asset_code.lower(), name.lower(), cls_obj.code])
    fp = hashlib.sha256(raw.encode()).hexdigest()

    # Exact fingerprint match → DUPLICATE
    existing_fp = (
        Asset.objects.filter(fingerprint=fp, site=site, is_deleted=False)
        .only("id", "name")
        .first()
    )
    if existing_fp:
        result["status"] = "DUPLICATE"
        result["existing_asset_id"] = str(existing_fp.pk)
        return result

    # Asset code match but different fingerprint → UPDATE_CANDIDATE
    existing_code = (
        Asset.objects.filter(asset_code=asset_code, site=site, is_deleted=False)
        .only("id")
        .first()
    )
    if existing_code:
        result["status"] = "UPDATE_CANDIDATE"
        result["existing_asset_id"] = str(existing_code.pk)
        return result

    # NEW
    return result


# ---------------------------------------------------------------------------
# Corrections (phase 2)
# ---------------------------------------------------------------------------

def apply_corrections(results_json: dict, corrections: list[dict], tenant, site) -> dict:
    """
    Apply field-level corrections to REJECTED rows and re-classify them.
    Non-REJECTED rows are left untouched.
    Returns the updated results_json.
    """
    cls_map: dict[str, AssetClassification] = {
        c.code: c
        for c in AssetClassification.objects.filter(tenant=tenant, is_active=True)
    }

    # Index rows for fast lookup
    rows_by_num = {r["row_number"]: r for r in results_json["rows"]}

    # Apply corrections
    for corr in corrections:
        row_num   = corr["row_number"]
        field     = corr["field"].strip().lower()
        new_value = str(corr["new_value"]).strip()
        row = rows_by_num.get(row_num)
        if row is None or row["status"] != "REJECTED":
            continue  # Only correct REJECTED rows
        if field == "asset_code":
            row["asset_code"] = new_value.upper()
        elif field == "name":
            row["name"] = new_value
        elif field == "classification_code":
            row["classification_code"] = new_value.upper()
        else:
            row.setdefault("custom_data", {})[field] = new_value

    # Rebuild seen_codes from non-REJECTED, non-BATCH_DUPLICATE rows
    seen_codes: dict[str, int] = {}
    for row in results_json["rows"]:
        if row["status"] not in ("REJECTED", "BATCH_DUPLICATE") and row["asset_code"]:
            seen_codes[row["asset_code"]] = row["row_number"]

    # Re-classify only the previously-REJECTED rows
    for row in results_json["rows"]:
        if row["status"] != "REJECTED":
            continue
        # Build synthetic row dict for re-classification
        synthetic = {
            "_row_number":       row["row_number"],
            "asset_code":        row.get("asset_code", ""),
            "name":              row.get("name", ""),
            "classification_code": row.get("classification_code", ""),
            **row.get("custom_data", {}),
        }
        new_result = _classify_single_row(synthetic, site, cls_map, seen_codes)
        row.update(new_result)
        if row["status"] not in ("REJECTED", "BATCH_DUPLICATE") and row["asset_code"]:
            seen_codes.setdefault(row["asset_code"], row["row_number"])

    return results_json


# ---------------------------------------------------------------------------
# Confirmation (phase 3)
# ---------------------------------------------------------------------------

@transaction.atomic
def confirm_import(job: BulkImportJob, decisions: list[dict], user) -> dict:
    """
    Apply create/update/skip decisions for each non-REJECTED row.
    Entire operation runs in a single transaction: crash → full rollback.
    Updates job.status → CONFIRMED on success.
    Returns {created, updated, skipped}.
    """
    decision_map = {int(d["row_number"]): d["action"] for d in decisions}
    created = updated = skipped = 0

    for row in job.results_json["rows"]:
        row_num = row["row_number"]
        status  = row["status"]

        # REJECTED / BATCH_DUPLICATE are always skipped
        if status in ("REJECTED", "BATCH_DUPLICATE"):
            skipped += 1
            continue

        action = decision_map.get(row_num, "skip")

        if action == "skip":
            skipped += 1
            continue

        cls = AssetClassification.objects.get(
            code=row["classification_code"], tenant=job.tenant
        )

        if action == "create":
            fp_raw = "|".join([
                str(job.site_id),
                row["asset_code"].lower(),
                row["name"].lower(),
                cls.code,
            ])
            fp = hashlib.sha256(fp_raw.encode()).hexdigest()
            asset = Asset.objects.create(
                site=job.site,
                asset_code=row["asset_code"],
                name=row["name"],
                classification=cls,
                fingerprint=fp,
            )
            asset.create_version(
                data=row.get("custom_data", {}),
                source=AssetVersion.ChangeSource.BULK_IMPORT,
                user=user,
            )
            created += 1

        elif action == "update":
            existing_id = row.get("existing_asset_id")
            if not existing_id:
                skipped += 1
                continue
            asset = Asset.objects.select_related("classification").get(pk=existing_id)
            asset.name           = row["name"]
            asset.classification = cls
            Asset.objects.filter(pk=asset.pk).update(
                name=row["name"], classification=cls
            )
            asset.create_version(
                data=row.get("custom_data", {}),
                source=AssetVersion.ChangeSource.BULK_IMPORT,
                user=user,
            )
            updated += 1

    job.status       = BulkImportJob.Status.CONFIRMED
    job.completed_at = timezone.now()
    BulkImportJob.objects.filter(pk=job.pk).update(
        status=job.status, completed_at=job.completed_at
    )
    return {"created": created, "updated": updated, "skipped": skipped}


# ---------------------------------------------------------------------------
# Export
# ---------------------------------------------------------------------------

def build_export(user, site_id: str | None, format_type: str) -> tuple[bytes, str, str]:
    """
    Build and return (file_bytes, content_type, filename) for an asset export.
    Columns: asset_code, name, classification_code, version_number,
             last_updated, + all custom_data keys (flattened).
    """
    from assets.views import _base_asset_queryset  # avoid circular import at module load

    qs = _base_asset_queryset(user).filter(is_deleted=False)
    if site_id:
        qs = qs.filter(site_id=site_id)

    # Collect rows and all custom_data field names
    rows_data: list[dict] = []
    extra_fields: list[str] = []
    extra_set: set[str] = set()

    for asset in qs.order_by("asset_code"):
        snapshot = asset.current_version.data_snapshot if asset.current_version else {}
        for key in snapshot:
            if key not in extra_set:
                extra_set.add(key)
                extra_fields.append(key)
        rows_data.append({
            "asset_code":          asset.asset_code,
            "name":                asset.name,
            "classification_code": asset.classification.code,
            "version_number":      asset.current_version.version_number if asset.current_version else "",
            "last_updated":        (
                asset.current_version.created_at.isoformat()
                if asset.current_version else ""
            ),
            **snapshot,
        })

    base_headers = ["asset_code", "name", "classification_code", "version_number", "last_updated"]
    all_headers  = base_headers + extra_fields

    if format_type == "xlsx":
        return _to_xlsx(rows_data, all_headers)
    else:
        return _to_csv(rows_data, all_headers)


def _to_xlsx(rows: list[dict], headers: list[str]) -> tuple[bytes, str, str]:
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Assets"
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h, "") for h in headers])

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", "assets_export.xlsx"


def _to_csv(rows: list[dict], headers: list[str]) -> tuple[bytes, str, str]:
    buf = io.StringIO()
    writer = csv.DictWriter(buf, fieldnames=headers, extrasaction="ignore")
    writer.writeheader()
    for row in rows:
        writer.writerow({h: row.get(h, "") for h in headers})
    return buf.getvalue().encode("utf-8"), "text/csv", "assets_export.csv"


# ---------------------------------------------------------------------------
# Preview summary helper (used by views)
# ---------------------------------------------------------------------------

def summarise(results_json: dict) -> dict:
    rows = results_json.get("rows", [])
    counts: dict[str, int] = {
        "new_count":             0,
        "update_count":          0,
        "duplicate_count":       0,
        "rejected_count":        0,
        "batch_duplicate_count": 0,
    }
    for row in rows:
        s = row.get("status", "")
        if s == "NEW":
            counts["new_count"] += 1
        elif s == "UPDATE_CANDIDATE":
            counts["update_count"] += 1
        elif s == "DUPLICATE":
            counts["duplicate_count"] += 1
        elif s == "REJECTED":
            counts["rejected_count"] += 1
        elif s == "BATCH_DUPLICATE":
            counts["batch_duplicate_count"] += 1
    counts["total"] = len(rows)
    return counts
